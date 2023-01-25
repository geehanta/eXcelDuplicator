import openpyxl

# Open the Excel file
wb = openpyxl.load_workbook('SCRIPT1.xlsx')

# Select the active sheet
ws = wb.active

# Create a new workbook and add a new sheet
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# Iterate through the rows and columns in the sheet
for row in ws.iter_rows():
    for cell in row:
        # Duplicate the cell six times
        for i in range(6):
            new_cell = new_ws.cell(row=cell.row, column=cell.column + i + 1)
            new_cell.value = cell.value

# Save the new workbook to a new file
new_wb.save('duplicated_example2.xlsx')